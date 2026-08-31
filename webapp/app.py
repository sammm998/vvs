"""Webbapp för mängdningsverktyget.

Ladda upp en CAD-exporterad PDF-ritning i webbläsaren, följ bearbetningen
och ladda ner den markerade PDF:en, mängdförteckningen (XLSX/CSV),
kodtabellen och rapporten.

Körning lokalt:
    uvicorn webapp.app:app --reload

Bearbetningen (OCR på A1-ritningar i 450 DPI) tar flera minuter och körs
därför i en bakgrundstråd per jobb; webbläsaren pollar statusendpointen.
Skalverifieringen kan inte vara interaktiv här – uppmätt skala redovisas i
resultatet och kan överstyras med skalfältet i formuläret.
"""

from __future__ import annotations

import logging
import os
import shutil
import threading
import time
import uuid
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse

from mangdning.config import Config, parse_vertical_heights
from mangdning.pipeline import run_pipeline

log = logging.getLogger("webapp")

app = FastAPI(title="Mängdning av VVS/VA-ritningar")

JOBS_DIR = Path(os.environ.get("JOBS_DIR", "jobs"))
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "100"))
JOB_TTL_SECONDS = int(os.environ.get("JOB_TTL_SECONDS", str(24 * 3600)))
# OCR är CPU-tungt – kör ett jobb i taget om inget annat anges
_worker_slots = threading.Semaphore(int(os.environ.get("WORKERS", "1")))

JOBS: dict[str, dict] = {}
_jobs_lock = threading.Lock()

STATIC_DIR = Path(__file__).parent / "static"


def _safe_pdf_name(filename: str | None) -> str:
    name = Path(filename or "ritning.pdf").name
    stem = "".join(c for c in Path(name).stem
                   if c.isalnum() or c in "._- ")[:80].strip() or "ritning"
    return f"{stem}.pdf"


def _zones_to_points(spec: str, pdf_bytes: bytes, page_index: int
                     ) -> list[tuple[float, float, float, float]]:
    """Områden som användaren ritat på förhandsvisningen kommer som andelar
    av bildens bredd/höjd (0-1) och räknas här om till PDF-punkter, så de är
    oberoende av vilken DPI förhandsvisningen renderades i."""
    import fitz

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        rect = doc[page_index].rect
    finally:
        doc.close()

    zones = []
    for part in spec.split(";"):
        part = part.strip()
        if not part:
            continue
        values = [float(v) for v in part.split(",")]
        if len(values) != 4:
            raise ValueError(f"Ogiltigt område '{part}'")
        fx0, fy0, fx1, fy1 = values
        zones.append((min(fx0, fx1) * rect.width, min(fy0, fy1) * rect.height,
                      max(fx0, fx1) * rect.width, max(fy0, fy1) * rect.height))
    return zones


def _cleanup_old_jobs() -> None:
    now = time.time()
    with _jobs_lock:
        expired = [jid for jid, job in JOBS.items()
                   if now - job["created"] > JOB_TTL_SECONDS]
        for jid in expired:
            JOBS.pop(jid, None)
            shutil.rmtree(JOBS_DIR / jid, ignore_errors=True)


def _run_job(job_id: str, cfg: Config, facit_path: Path | None) -> None:
    job = JOBS[job_id]
    job_dir = JOBS_DIR / job_id
    with _worker_slots:
        job["status"] = "running"
        try:
            outputs = run_pipeline(
                job_dir / job["input_name"], cfg, job_dir / "out",
                facit=facit_path,
                on_stage=lambda sid, label: job.update(stage=label))
            files = {
                "markerad_pdf": outputs.annotated_pdf,
                "mangder_xlsx": outputs.quantities_xlsx,
                "mangder_csv": outputs.quantities_csv,
                "koder_csv": outputs.code_table_csv,
                "rapport_txt": outputs.report_txt,
            }
            if outputs.preview_png:
                files["forhandsvisning_png"] = outputs.preview_png
            if outputs.validation_txt:
                files["validering_txt"] = outputs.validation_txt
            job["files"] = {key: p.name for key, p in files.items()}
            job["summary"] = outputs.summary
            job["rows"] = _table_rows(outputs.quantities_xlsx)
            job["status"] = "done"
        except Exception as exc:
            log.exception("Jobb %s misslyckades", job_id)
            job["status"] = "error"
            job["error"] = str(exc)


def _table_rows(xlsx_path: Path, limit: int = 3000) -> list[dict]:
    """Mängdraderna som JSON, så resultatsidan kan visa uträkningen direkt
    i webbläsaren i stället för att kräva nedladdning av Excel-filen."""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Mängdförteckning"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            "subject": r[2], "color": r[5], "langd": r[7], "lager": r[9],
            "antal_vs": r[10], "total_vh": r[12],
            "antal": r[18], "kalla": r[19],
        })
        if len(rows) >= limit:
            break
    return rows


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return (STATIC_DIR / "index.html").read_text(encoding="utf-8")


@app.post("/api/jobs")
async def create_job(
    file: UploadFile = File(...),
    facit: UploadFile | None = File(default=None),
    scale: str = Form(default=""),
    pipe_width: str = Form(default=""),
    vertikalhojd: str = Form(default=""),
    page: int = Form(default=1),
    dpi: str = Form(default=""),
    ocr: str = Form(default="auto"),  # auto | force | off
    exclude_zones: str = Form(default=""),  # "fx0,fy0,fx1,fy1;..." 0-1
):
    _cleanup_old_jobs()
    data = await file.read()
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(413, f"Filen är större än {MAX_UPLOAD_MB} MB")
    if not data.startswith(b"%PDF"):
        raise HTTPException(400, "Filen verkar inte vara en PDF")

    cfg = Config()
    try:
        if scale.strip():
            cfg.scale = scale.strip()
        if pipe_width.strip():
            cfg.pipe_width = float(pipe_width.replace(",", "."))
        if vertikalhojd.strip():
            cfg.vertical_heights.update(parse_vertical_heights(vertikalhojd))
        if dpi.strip():
            cfg.dpi = max(72, min(int(dpi), 900))
        cfg.page = max(0, page - 1)  # formuläret är 1-baserat
        if ocr == "force":
            cfg.force_ocr = True
        elif ocr == "off":
            cfg.skip_ocr = True
        elif ocr != "auto":
            raise ValueError(f"ocr måste vara auto/force/off, inte '{ocr}'")
        if exclude_zones.strip():
            cfg.exclude_zones.extend(
                _zones_to_points(exclude_zones, data, cfg.page))
    except ValueError as exc:
        raise HTTPException(400, f"Ogiltig parameter: {exc}")

    job_id = uuid.uuid4().hex[:12]
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_name = _safe_pdf_name(file.filename)
    (job_dir / input_name).write_bytes(data)

    facit_path: Path | None = None
    if facit is not None and facit.filename:
        facit_data = await facit.read()
        if len(facit_data) > 20 * 1024 * 1024:
            raise HTTPException(413, "Facit-filen är för stor")
        facit_path = job_dir / "facit.csv"
        facit_path.write_bytes(facit_data)

    JOBS[job_id] = {
        "status": "queued", "stage": "Väntar på ledig plats i kön",
        "created": time.time(), "input_name": input_name,
        "files": {}, "summary": None, "error": None, "rows": [],
    }
    threading.Thread(target=_run_job, args=(job_id, cfg, facit_path),
                     daemon=True).start()
    return {"job_id": job_id}


@app.get("/api/jobs/{job_id}")
def job_status(job_id: str):
    job = JOBS.get(job_id)
    if job is None:
        raise HTTPException(404, "Okänt jobb (kan ha rensats efter 24 h)")
    return {
        "status": job["status"],
        "stage": job.get("stage"),
        "error": job.get("error"),
        "files": job["files"],
        "summary": job["summary"],
        "rows": job.get("rows") or [],
    }


@app.get("/api/jobs/{job_id}/files/{name}")
def job_file(job_id: str, name: str):
    job = JOBS.get(job_id)
    if job is None:
        raise HTTPException(404, "Okänt jobb")
    if name not in job["files"].values():
        raise HTTPException(404, "Okänd fil")  # bara registrerade utdatafiler
    path = JOBS_DIR / job_id / "out" / name
    if not path.exists():
        raise HTTPException(404, "Filen saknas på disk")
    media = {"pdf": "application/pdf", "csv": "text/csv", "png": "image/png",
             "txt": "text/plain; charset=utf-8",
             "xlsx": ("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet")}.get(path.suffix.lstrip("."))
    if path.suffix.lower() == ".png":
        return FileResponse(path, media_type="image/png")
    return FileResponse(path, media_type=media, filename=name)


@app.get("/health")
def health():
    return {"status": "ok"}
