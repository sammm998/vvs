"""Output-skrivare: mängdförteckning (XLSX + CSV), kodtabell (CSV) och
körningsrapport med kända begränsningar.

Mängdförteckningens huvudflik använder facit-filens kolumnstruktur
(Version, Document, Subject, ..., Kontroll_Vs) så att resultatet är direkt
strukturellt jämförbart med en professionell mängdning, plus egna
spårbarhetskolumner (Antal, Källa) sist.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from .models import CodeHit, PipeChain, ScaleResult
from .quantify import QuantityResult

log = logging.getLogger(__name__)

# Exakt kolumnuppsättning ur facit-exporten, så att resultatet kan läggas
# sida vid sida med en professionell mängdning.
FACIT_COLUMNS = [
    "Version", "Document", "Subject", "Sorterings_siffror", "Sidetikett",
    "Color", "Kommentarer", "Längd", "unit", "Lager", "Antal_VS",
    "Vertikal_höjd_VS", "Total_vertikalhöjd_VS", "unit2",
    "Kapitel_i_Sektionsdata", "Exportval_Vs", "Byggdelsnummer",
    "Kontroll_Vs",
]
# Egna spårbarhetskolumner sist – finns inte i facit men behövs för att
# kunna gå tillbaka till rätt rörsträcka/kod i den markerade PDF:en.
EXTRA_COLUMNS = ["Antal", "Källa", "Osäkerhet"]

VERSION = "1.0"


def _num(value: float | None) -> str:
    """Tal med decimalkomma, som i facit."""
    return "" if value is None else f"{value:.1f}".replace(".", ",")


def _facit_row(row) -> list:
    is_vertical = row.antal_vs is not None
    if is_vertical:
        kommentar = str(row.antal_vs)
        exportval = "Total vertikal höjd VS"
    elif row.langd_m is not None:
        kommentar = f"{_num(row.langd_m)} m"
        exportval = "Längd"
    else:
        kommentar = ""
        exportval = ""
    return [
        VERSION,                  # Version
        row.document,             # Document
        row.subject,              # Subject
        "",                       # Sorterings_siffror
        row.sidetikett,           # Sidetikett
        row.color,                # Color
        kommentar,                # Kommentarer
        row.langd_m if row.langd_m is not None else "",   # Längd
        "m" if row.langd_m is not None else "",           # unit
        row.lager,                # Lager
        row.antal_vs if row.antal_vs is not None else "",
        row.vertikal_hojd_m if row.vertikal_hojd_m is not None else "",
        row.total_vertikalhojd_m if row.total_vertikalhojd_m is not None else "",
        "m" if row.total_vertikalhojd_m is not None else "",  # unit2
        "",                       # Kapitel_i_Sektionsdata
        exportval,                # Exportval_Vs
        "",                       # Byggdelsnummer
        "Avmarkerad",             # Kontroll_Vs
        row.antal if row.antal is not None else "",       # Antal (st)
        row.kalla,                # Källa
        row.kommentar,            # Osäkerhet (våra flaggor)
    ]


def write_quantities_csv(result: QuantityResult, path: str | Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(FACIT_COLUMNS + EXTRA_COLUMNS)
        for row in result.rows:
            writer.writerow(_facit_row(row))
    log.info("Mängdförteckning (CSV) sparad: %s", path)


def write_quantities_xlsx(result: QuantityResult, path: str | Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # --- Flik 1: detaljerad radnivå (huvudleveransen) ---
    ws = wb.active
    ws.title = "Mängdförteckning"
    ws.append(FACIT_COLUMNS + EXTRA_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in result.rows:
        ws.append(_facit_row(row))
        color_cell = ws.cell(row=ws.max_row, column=6)
        try:
            argb = "FF" + row.color.lstrip("#")
            color_cell.fill = PatternFill("solid", fgColor=argb)
        except Exception:
            pass

    # --- Flik 2: aggregerad summering per kod ---
    ws2 = wb.create_sheet("Summering per kod")
    ws2.append(["Subject", "Lager", "Color", "Antal sträckor",
                "Total längd (m)", "Antal (st)", "Antal_VS",
                "Total vertikalhöjd (m)", "Flaggor"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for agg in result.aggregates:
        ws2.append([
            agg.subject, agg.lager, agg.color, agg.n_strackor,
            agg.total_langd_m if agg.n_strackor else "",
            agg.n_punkter or "", agg.antal_vs or "",
            agg.total_vertikalhojd_m if agg.antal_vs else "",
            "; ".join(sorted(agg.flaggor)),
        ])

    # --- Flik 3: skalfaktor och hur den räknades fram ---
    ws3 = wb.create_sheet("Skala")
    s = result.scale
    ws3.append(["Parameter", "Värde"])
    ws3["A1"].font = ws3["B1"].font = Font(bold=True)
    ws3.append(["Metod", s.method])
    ws3.append(["Skaltext", s.scale_text or "-"])
    ws3.append(["Punkter per meter (använd)",
                round(s.pts_per_meter, 4) if s.known else "OKÄND"])
    ws3.append(["Titelblock (pt/m)",
                round(s.title_pts_per_meter, 4) if s.title_pts_per_meter else "-"])
    ws3.append(["Skalstock (pt/m)",
                round(s.bar_pts_per_meter, 4) if s.bar_pts_per_meter else "-"])
    for w in s.warnings:
        ws3.append(["VARNING", w])

    # --- Flik 4: varningar/osäkerhetsflaggor ---
    ws4 = wb.create_sheet("Varningar")
    ws4.append(["Varning"])
    ws4["A1"].font = Font(bold=True)
    for w in result.warnings:
        ws4.append([w])
    for c in result.count_checks:
        ws4.append([c])

    wb.save(str(path))
    log.info("Mängdförteckning (XLSX) sparad: %s", path)


def write_code_table(codes: list[CodeHit], path: str | Path) -> None:
    """Kodtabell: kod, position, OCR-confidence, kopplad rörsträcka."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["id", "kod", "dimension", "full_kod", "antal(N)",
                         "x", "y", "ocr_confidence", "råa_ocr_träffar",
                         "kopplad_rörsträcka", "kopplingsmetod",
                         "exkluderad", "exkluderingsorsak"])
        for c in codes:
            cx, cy = c.bbox.center
            writer.writerow([
                c.id, c.raw_text, c.dimension or "", c.full_code, c.count,
                round(cx, 1), round(cy, 1), round(c.conf, 1),
                c.raw_cluster_size,
                c.linked_chain if c.linked_chain is not None else "",
                c.link_method or "",
                "ja" if c.excluded else "nej", c.excluded_reason or "",
            ])
    log.info("Kodtabell sparad: %s", path)


def write_run_report(path: str | Path, *, input_pdf: str,
                     codes: list[CodeHit], chains: list[PipeChain],
                     scale: ScaleResult, result: QuantityResult,
                     pipe_width: float, pipe_color, text_info,
                     histogram_text: str) -> None:
    """Logg/rapport över körningen och dess kända begränsningar."""
    active_chains = [c for c in chains if not c.excluded]
    linked_chains = [c for c in active_chains if c.linked_codes]
    active_codes = [c for c in codes if not c.excluded]
    linked_codes = [c for c in active_codes if c.linked_chain is not None]

    lines = [
        "KÖRNINGSRAPPORT – mängdning av VVS/VA-ritning",
        "=" * 60,
        f"Indata: {input_pdf}",
        "",
        "-- Textlager --",
        f"Riktiga PDF-textord: {text_info.native_words}",
        f"Vektor-paths: {text_info.vector_paths}",
        f"OCR användes: {'ja (vektoriserad text)' if text_info.use_ocr else 'nej'}",
        "",
        "-- Skala --",
        f"Metod: {scale.method}   Skaltext: {scale.scale_text or '-'}",
        f"Punkter per meter: "
        f"{f'{scale.pts_per_meter:.4f}' if scale.known else 'OKÄND'}",
        "",
        "-- Rördetektering --",
        f"Valt rörkluster: bredd={pipe_width:.2f} pt, färg={pipe_color}",
        f"Rörsträckor: {len(active_chains)} "
        f"({len(chains) - len(active_chains)} exkluderade som ram/rutnät)",
        f"  varav kopplade till kod: {len(linked_chains)}",
        f"  varav OKOPPLADE (verifiera manuellt): "
        f"{len(active_chains) - len(linked_chains)}",
        "",
        "-- Koder --",
        f"Godkända koder: {len(active_codes)} "
        f"({len(codes) - len(active_codes)} exkluderade via zoner)",
        f"  varav kopplade till rör: {len(linked_codes)}",
        f"  varav EJ kopplade till rör (punktkomponenter eller "
        f"legend/tabelltext – verifiera manuellt): "
        f"{len(active_codes) - len(linked_codes)}",
        "",
        "-- Mängdförteckning --",
        f"Rader totalt: {len(result.rows)}",
        f"Unika koder: {len(result.aggregates)}",
        "",
        "-- Antals-sanity-check mot legend --",
        *(result.count_checks or ["(inga antalsuppgifter lästa ur legenden)"]),
        "",
        "-- Varningar / kända begränsningar --",
        *[f"* {w}" for w in result.warnings],
        "* Rördetekteringens linjebredd/färg-tröskel är filspecifik och kan "
        "behöva kalibreras om för andra ritningsmallar (--calibrate, "
        "--pipe-width).",
        "* Med originalfilen (DWG) i stället för PDF skulle text och skala "
        "kunna läsas exakt utan OCR/gissning – möjlig framtida importväg.",
        "",
        "-- Kalibreringsunderlag --",
        histogram_text,
    ]
    Path(path).write_text("\n".join(lines), encoding="utf-8")
    log.info("Körningsrapport sparad: %s", path)
