"""Del D punkt 9 – valideringsmodul.

Jämför verktygets resultat mot en facit-fil (mängdförteckning exporterad
från ett professionellt mängdningsverktyg, med kolumnerna Version, Document,
Subject, ..., Längd, Lager, Antal_VS, ...). Rapporterar per kod: antal
rader, summerad längd och differens i procent. Så mäts och förbättras
pipelinens träffsäkerhet iterativt – inte bara genom visuell koll.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, field
from pathlib import Path

from .quantify import QuantityResult

log = logging.getLogger(__name__)


@dataclass
class FacitEntry:
    subject: str
    n_rows: int = 0
    n_length_rows: int = 0
    total_length: float = 0.0
    n_points: int = 0
    n_vertical_rows: int = 0
    antal_vs: int = 0
    lager: str = ""


@dataclass
class ValidationReport:
    lines: list[str] = field(default_factory=list)
    n_codes_compared: int = 0
    n_codes_ok: int = 0

    def text(self) -> str:
        return "\n".join(self.lines)


def _cell(row: dict, key: str) -> str:
    """Cellvärde som text – XLSX ger tal/None, CSV ger strängar."""
    value = row.get(key)
    return "" if value is None else str(value).strip()


def _to_float(value: str) -> float | None:
    value = (value or "").strip().replace(" ", "").replace(" ", "")
    if not value:
        return None
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return None


def _to_int(value: str) -> int | None:
    f = _to_float(value)
    return int(f) if f is not None else None


def _read_facit_rows(path: Path) -> list[dict]:
    """Läs facit-raderna ur XLSX eller CSV till dict per rad."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Facit-filen är tom: {path}")
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:]]

    raw = path.read_text(encoding="utf-8-sig")
    if not raw.strip():
        raise ValueError(f"Facit-filen är tom: {path}")
    header_line = raw.splitlines()[0]
    delimiter = ";" if header_line.count(";") >= header_line.count(",") else ","
    return list(csv.DictReader(raw.splitlines(), delimiter=delimiter))


def read_facit(path: str | Path) -> dict[str, FacitEntry]:
    """Läs en facit-mängdförteckning (XLSX eller CSV, decimalkomma stöds)."""
    path = Path(path)
    rows = _read_facit_rows(path)
    if rows and "Subject" not in rows[0]:
        raise ValueError(
            f"Facit-filen saknar 'Subject'-kolumn "
            f"(kolumner: {list(rows[0])})")

    entries: dict[str, FacitEntry] = {}
    for row in rows:
        subject = _cell(row, "Subject")
        if not subject:
            continue
        entry = entries.setdefault(subject, FacitEntry(subject=subject))
        entry.n_rows += 1
        entry.lager = entry.lager or _cell(row, "Lager")
        length = _to_float(_cell(row, "Längd") or _cell(row, "Langd"))
        antal_vs = _to_int(_cell(row, "Antal_VS"))
        if length is not None:
            entry.n_length_rows += 1
            entry.total_length += length
        elif antal_vs:
            entry.n_vertical_rows += 1
        else:
            entry.n_points += 1
        if antal_vs:
            entry.antal_vs += antal_vs
    log.info("Facit inläst: %d rader, %d unika koder",
             sum(e.n_rows for e in entries.values()), len(entries))
    return entries


def validate_against_facit(result: QuantityResult,
                           facit: dict[str, FacitEntry],
                           length_tol_pct: float = 5.0) -> ValidationReport:
    """Avvikelserapport per kod: t.ex.
    "S3-R8-75: facit 27 sträckor / 118,3 m totalt,
     vårt resultat 24 sträckor / 109,6 m totalt, -7 %"."""
    report = ValidationReport()
    ours = {a.subject: a for a in result.aggregates}

    report.lines.append("VALIDERING MOT FACIT")
    report.lines.append("=" * 60)

    for subject in sorted(facit):
        f = facit[subject]
        o = ours.get(subject)
        report.n_codes_compared += 1
        if o is None:
            report.lines.append(
                f"{subject}: facit {f.n_length_rows} sträckor / "
                f"{_fmt(f.total_length)} m totalt ({f.n_points} punkter), "
                f"vårt resultat SAKNAS helt")
            continue

        parts = [f"{subject}: facit {f.n_length_rows} sträckor / "
                 f"{_fmt(f.total_length)} m totalt, "
                 f"vårt resultat {o.n_strackor} sträckor / "
                 f"{_fmt(o.total_langd_m)} m totalt"]
        ok = True
        if f.total_length > 0:
            diff_pct = (o.total_langd_m - f.total_length) / f.total_length * 100.0
            parts.append(f"{diff_pct:+.0f} %")
            if abs(diff_pct) > length_tol_pct:
                ok = False
        elif o.n_strackor != f.n_length_rows:
            ok = False
        if f.n_points or o.n_punkter:
            parts.append(f"punkter: facit {f.n_points} / vårt {o.n_punkter}")
            if f.n_points != o.n_punkter:
                ok = False
        if f.antal_vs or o.antal_vs:
            parts.append(f"vertikaler: facit {f.antal_vs} / vårt {o.antal_vs}")
        if ok:
            report.n_codes_ok += 1
        report.lines.append(", ".join(parts) + ("" if ok else "  [AVVIKELSE]"))

    only_ours = sorted(set(ours) - set(facit))
    if only_ours:
        report.lines.append("")
        report.lines.append("Koder hos oss som saknas i facit "
                            "(möjliga OCR-feltolkningar eller legendtext):")
        for subject in only_ours:
            o = ours[subject]
            report.lines.append(
                f"  {subject}: {o.n_strackor} sträckor / "
                f"{_fmt(o.total_langd_m)} m, {o.n_punkter} punkter")

    report.lines.append("")
    report.lines.append(
        f"Sammanfattning: {report.n_codes_ok}/{report.n_codes_compared} "
        f"koder inom tolerans (±{length_tol_pct:.0f} % längd, exakt antal "
        f"punkter)")
    return report


def _fmt(v: float) -> str:
    return f"{v:.1f}".replace(".", ",")
